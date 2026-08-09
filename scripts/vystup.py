# -*- coding: utf-8 -*-
"""
Sestaví sešit připravený k překopírování do modelu databaze.xlsx.

Kromě listu Přehled vzniknou tři listy, jejichž rozvržení je totožné
s listem „IMPORT HLASOVACÍCH LÍSTKŮ" v modelu — stejná písmena sloupců
i stejná čísla řádků. Obsluha označí žlutou oblast, přepne se do modelu
a vloží ji na tutéž adresu. Nic se nepřerovnává ručně.

Rozsahy v modelu jsou pevné (VLOOKUP přes $D$3:$L$8 a $D$21:$U$25,
COUNTIF přes $E$12:$E$16), proto má každý blok právě pět řádků.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import spolecne as s

PISMO = "Arial"
NADPIS = Font(name=PISMO, size=14, bold=True)
POKYN = Font(name=PISMO, size=10, bold=True, color="C00000")
HLAVICKA = Font(name=PISMO, size=9, bold=True, color="FFFFFF")
BEZNE = Font(name=PISMO, size=10)
SLABE = Font(name=PISMO, size=9, color="666666")
VYPLN = PatternFill("solid", fgColor="1F3864")
KOPIE = PatternFill("solid", fgColor="FFF2CC")     # oblast určená ke zkopírování
RAMECEK = Border(bottom=Side(style="thin", color="BFBFBF"))
ZELENA = Font(name=PISMO, size=10, color="0D7A44", bold=True)
CERVENA = Font(name=PISMO, size=10, color="C00000")
DATUM = "yyyy-mm-dd hh:mm:ss"

# Počet řádků, na které jsou v modelu nastavené vzorce.
KAPACITA = 5

# --------------------------------------------------------------------------
# Rozvržení cílového listu. Poslední položka trojice = hlavička, kterou
# posílá formulář; @timestamp znamená časovou značku.
# --------------------------------------------------------------------------
ROZVRZENI = {
    "navrhy": {
        "list": "1 · Návrhy",
        "nadpis": "TÉMATA - NÁVRHY (formulář č. 1)",
        "radek_hlavicky": 3,
        "prvni_radek": 4,
        "poradi": True,
        "poradi_kopirovat": False,
        "sloupce": [
            ("B", "Časová značka", "@timestamp"),
            ("C", "Uživatelské jméno", "Email Address"),
            ("D", "Jméno a příjmení", "Jméno a příjmení"),
            ("E", "AKCIE - ZAŘAZENÍ NOVÉHO TÉMATU č. 1", "AKCIE - ZAŘAZENÍ NOVÉHO TÉMATU č. 1"),
            ("F", "AKCIE - ZAŘAZENÍ NOVÉHO TÉMATU č. 2", "AKCIE - ZAŘAZENÍ NOVÉHO TÉMATU č. 2"),
            ("G", "AKCIE - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 1", "AKCIE - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 1"),
            ("H", "AKCIE - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 2", "AKCIE - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 2"),
            ("I", "DLUHOPISY - ZAŘAZENÍ NOVÉHO TÉMATU č. 1", "DLUHOPISY - ZAŘAZENÍ NOVÉHO TÉMATU č. 1"),
            ("J", "DLUHOPISY - ZAŘAZENÍ NOVÉHO TÉMATU č. 2", "DLUHOPISY - ZAŘAZENÍ NOVÉHO TÉMATU č. 2"),
            ("K", "DLUHOPISY - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 1", "DLUHOPISY - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 1"),
            ("L", "DLUHOPISY - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 2", "DLUHOPISY - VYŘAZENÍ STÁVAJÍCÍHO TÉMATU č. 2"),
        ],
    },
    "priority": {
        "list": "2 · Priority",
        "nadpis": "TÉMATA - PRIORITY (formulář č. 2)",
        "radek_hlavicky": 11,
        "prvni_radek": 12,
        # Model počítá odevzdané lístky přes COUNT($A$12:$A$16),
        # proto se pořadové číslo musí kopírovat s sebou.
        "poradi": True,
        "poradi_kopirovat": True,
        "sloupce": [
            ("B", "Časová značka", "@timestamp"),
            ("C", "Uživatelské jméno", "Email Address"),
            ("D", "Jméno a příjmení", "Jméno a příjmení"),
            ("E", "Priorita 1", "AKCIE - Priorita 1"),
            ("F", "Priorita 2", "AKCIE - Priorita 2"),
            ("G", "Priorita 3", "AKCIE - Priorita 3"),
            ("H", "Priorita 4", "AKCIE - Priorita 4"),
            ("I", "Priorita 5", "AKCIE - Priorita 5"),
            ("J", "Priorita 1", "DLUHOPISY - Priorita 1"),
            ("K", "Priorita 2", "DLUHOPISY - Priorita 2"),
            ("L", "Priorita 3", "DLUHOPISY - Priorita 3"),
            ("M", "Priorita 4", "DLUHOPISY - Priorita 4"),
            ("N", "Priorita 5", "DLUHOPISY - Priorita 5"),
        ],
    },
    "hlavni": {
        "list": "3 · Hlavní část",
        "nadpis": "HLAVNÍ ČÁST (formulář č. 3)",
        "radek_hlavicky": 20,
        "prvni_radek": 21,
        "poradi": True,
        "poradi_kopirovat": False,
        "sloupce": [
            ("B", "Časová značka", "@timestamp"),
            ("C", "Uživatelské jméno", "Email Address"),
            ("D", "Jméno a příjmení", "Jméno a příjmení"),
            ("E", "TŘÍDY AKTIV [Peněžní trh]", "TŘÍDY AKTIV [Peněžní trh]"),
            ("F", "TŘÍDY AKTIV [Akcie]", "TŘÍDY AKTIV [Akcie]"),
            ("G", "TŘÍDY AKTIV [Dluhopisy]", "TŘÍDY AKTIV [Dluhopisy]"),
            ("H", "TŘÍDY AKTIV [Alternativy]", "TŘÍDY AKTIV [Alternativy]"),
            ("I", "AKCIE [Americké akcie]", "AKCIE [Americké akcie]"),
            ("J", "AKCIE [Evropské akcie]", "AKCIE [Evropské akcie]"),
            ("K", "AKCIE [Akcie rozvíjejících se trhů]", "AKCIE [Akcie rozvíjejících se trhů]"),
            ("L", "AKCIE [Témata / ostatní akciové pozice]", "AKCIE [Témata / ostatní akciové pozice]"),
            ("M", "DLUHOPISY [České státní dluhopisy]", "DLUHOPISY [České státní dluhopisy]"),
            ("N", "DLUHOPISY [Americké státní dluhopisy]", "DLUHOPISY [Americké státní dluhopisy]"),
            ("O", "DLUHOPISY [Evropské státní dluhopisy]", "DLUHOPISY [Evropské státní dluhopisy]"),
            ("P", "DLUHOPISY [Korporátní dluhopisy - investiční stupeň]", "DLUHOPISY [Korpo. dluhopisy - investiční stupeň]"),
            ("Q", "DLUHOPISY [Korporátní dluhopisy - spekulativní stupeň]", "DLUHOPISY [Korpo. dluhopisy - spekulativní stupeň]"),
            ("R", "DLUHOPISY [Ostatní dluhopisové pozice]", "DLUHOPISY [Ostatní dluhopisové pozice]"),
            ("S", "HLASOVACÍ LÍSTEK - FX [EUR/CZK]", "FX [EUR/CZK]"),
            ("T", "HLASOVACÍ LÍSTEK - FX [USD/CZK]", "FX [USD/CZK]"),
            ("U", "HLASOVACÍ LÍSTEK - FX [EUR/USD]", "FX [EUR/USD]"),
        ],
    },
}

_JEN_ZNAMENKA = re.compile(r"^[+\-\s]+$")


def normalizuj(hodnota) -> str:
    """
    Matice převádí hodnocení tabulkou +++ / ++ / + / N / - / -- / ---
    (bez mezer). Formulář může posílat „+ +“, proto mezery odstraníme —
    jinak by VLOOKUP v matici nenašel shodu a hlas by tiše propadl.
    """
    text = "" if hodnota is None else str(hodnota).strip()
    if text and _JEN_ZNAMENKA.match(text):
        return text.replace(" ", "")
    return text


def _cas(zaznam: dict):
    """Časová značka z lístku; když chybí nebo je nečitelná, čas přijetí."""
    mapa = dict(zaznam.get("pary", []))
    syrovy = (mapa.get("Timestamp") or "").strip()
    for tvar in ("%m/%d/%Y %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(syrovy, tvar)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(zaznam["prijato"].replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _zaznamy(slozka: Path, typ: str) -> list[dict]:
    f = slozka / f"{typ}.json"
    return json.loads(f.read_text(encoding="utf-8")) if f.exists() else []


def _list_prehled(wb: Workbook, kolo: dict) -> None:
    ws = wb.active
    ws.title = "Přehled"
    ws["A1"] = f"{kolo['nazev']} ({kolo['id']})"
    ws["A1"].font = NADPIS
    ws["A2"] = (f"Otevřeno {s.lidsky(kolo['otevreno'])}"
                + (f" · uzavřeno {s.lidsky(kolo['uzavreno'])}" if kolo.get("uzavreno")
                   else " · sběr stále běží")
                + (f" · termín {kolo['termin']}" if kolo.get("termin") else ""))
    ws["A2"].font = SLABE

    formulare = kolo["formulare"]
    zahlavi = ["Hlasující"] + [s.NAZVY_TYPU[t] for t in formulare]
    for i, text in enumerate(zahlavi, start=1):
        b = ws.cell(row=4, column=i, value=text)
        b.font, b.fill = HLAVICKA, VYPLN
        b.alignment = Alignment(horizontal="left" if i == 1 else "center", vertical="center")

    prvni = 5
    for r, h in enumerate(kolo["hlasujici"], start=prvni):
        b = ws.cell(row=r, column=1, value=h["jmeno"])
        b.font, b.border = BEZNE, RAMECEK
        for i, typ in enumerate(formulare, start=2):
            kdy = h.get("odevzdal", {}).get(typ)
            c = ws.cell(row=r, column=i, value="odevzdal" if kdy else "chybí")
            c.font = ZELENA if kdy else CERVENA
            c.alignment = Alignment(horizontal="center")
            c.border = RAMECEK
    posledni = prvni + len(kolo["hlasujici"]) - 1

    r = posledni + 1
    b = ws.cell(row=r, column=1, value="Odevzdáno")
    b.font = Font(name=PISMO, size=10, bold=True)
    pocet = len(kolo["hlasujici"])
    for i in range(2, len(formulare) + 2):
        p = get_column_letter(i)
        c = ws.cell(row=r, column=i,
                    value=f'=COUNTIF({p}{prvni}:{p}{posledni},"odevzdal")&" / {pocet}"')
        c.font = Font(name=PISMO, size=10, bold=True)
        c.alignment = Alignment(horizontal="center")

    ws.cell(row=r + 2, column=1, value="Jak přenést do modelu:").font = Font(name=PISMO, size=10, bold=True)
    pokyny = [
        "Každý další list odpovídá jednomu bloku listu „IMPORT HLASOVACÍCH LÍSTKŮ“ v modelu.",
        "Sloupce i čísla řádků jsou stejné, oblast se tedy vkládá na tutéž adresu, ze které se kopíruje.",
        "Žlutě podbarvenou oblast označte, zkopírujte a v modelu vložte jako hodnoty (Ctrl+Shift+V).",
        "Pořadí řádků nehraje roli — model hledá hlasující podle jména ve sloupci D.",
    ]
    for i, radek in enumerate(pokyny, start=1):
        ws.cell(row=r + 2 + i, column=1, value=f"{i}. {radek}").font = SLABE

    ws.column_dimensions["A"].width = 36
    for i in range(2, len(formulare) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "A5"


def _list_bloku(wb: Workbook, kolo: dict, slozka: Path, typ: str) -> list[str]:
    """Vytvoří list v rozvržení modelu. Vrací seznam varování."""
    rozvrh = ROZVRZENI[typ]
    ws = wb.create_sheet(rozvrh["list"])
    varovani: list[str] = []

    vsechny = sorted(_zaznamy(slozka, typ), key=lambda z: z["i"])
    zaznamy = vsechny[:KAPACITA]
    if len(vsechny) > KAPACITA:
        varovani.append(
            f"{s.NAZVY_TYPU[typ]}: lístků je {len(vsechny)}, ale vzorce v modelu počítají "
            f"s {KAPACITA}. Přebývající se do bloku nevešly — rozšiřte rozsahy v modelu.")

    hr, pr = rozvrh["radek_hlavicky"], rozvrh["prvni_radek"]
    posledni = pr + KAPACITA - 1
    prvni_sl = "A" if rozvrh["poradi_kopirovat"] else rozvrh["sloupce"][0][0]
    posledni_sl = rozvrh["sloupce"][-1][0]
    oblast = f"{prvni_sl}{pr}:{posledni_sl}{posledni}"

    ws["A1"] = rozvrh["nadpis"]
    ws["A1"].font = NADPIS
    ws.cell(row=hr - 1, column=1,
            value=f"Zkopírujte oblast {oblast} a v modelu ji vložte do listu "
                  f"„IMPORT HLASOVACÍCH LÍSTKŮ“ na adresu {prvni_sl}{pr} (vložit jako hodnoty).").font = POKYN

    for pismeno, nazev, _ in rozvrh["sloupce"]:
        b = ws[f"{pismeno}{hr}"]
        b.value = nazev
        b.font, b.fill = HLAVICKA, VYPLN
        b.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[hr].height = 40

    prirazene = set()
    for poradi, z in enumerate(zaznamy, start=1):
        r = pr + poradi - 1
        mapa = {h: v for h, v in z["pary"]}
        a = ws.cell(row=r, column=1, value=poradi)
        a.font = BEZNE
        a.alignment = Alignment(horizontal="center")
        if rozvrh["poradi_kopirovat"]:
            a.fill = KOPIE
        for pismeno, _, klic in rozvrh["sloupce"]:
            b = ws[f"{pismeno}{r}"]
            b.font, b.fill = BEZNE, KOPIE
            if klic == "@timestamp":
                b.value = _cas(z)
                b.number_format = DATUM
                b.alignment = Alignment(horizontal="left")
            elif klic in mapa:
                prirazene.add(klic)
                b.value = normalizuj(mapa[klic])

    # prázdné řádky do plné kapacity, ať oblast sedí na rozsah v modelu
    for r in range(pr + len(zaznamy), posledni + 1):
        for pismeno, _, _ in rozvrh["sloupce"]:
            ws[f"{pismeno}{r}"].fill = KOPIE
        if rozvrh["poradi_kopirovat"]:
            ws.cell(row=r, column=1).fill = KOPIE

    vsechny_klice = {h for z in zaznamy for h, _ in z["pary"]}
    nezarazene = sorted(vsechny_klice - prirazene - {"Timestamp"})
    ocekavane = {k for _, _, k in rozvrh["sloupce"] if k != "@timestamp"}
    chybejici = sorted(ocekavane - vsechny_klice) if zaznamy else []

    r = posledni + 2
    if nezarazene:
        varovani.append(f"{s.NAZVY_TYPU[typ]}: formulář posílá sloupce, pro které v modelu "
                        f"není místo: {', '.join(nezarazene)}")
        ws.cell(row=r, column=1,
                value="Pozor – tyto sloupce z formuláře se do modelu nevešly: "
                      + ", ".join(nezarazene)).font = CERVENA
        r += 1
    if chybejici:
        ws.cell(row=r, column=1,
                value="Model počítá se sloupci, které formulář nesbírá (zůstanou prázdné): "
                      + ", ".join(chybejici)).font = SLABE
        r += 1
    ws.cell(row=r, column=1,
            value=f"Kolo {kolo['id']} · lístků v bloku {len(zaznamy)} z {KAPACITA} "
                  f"· hodnocení jsou zapsána bez mezer (++ nikoli + +), jak je čte matice.").font = SLABE

    ws.column_dimensions["A"].width = 5
    for pismeno, nazev, klic in rozvrh["sloupce"]:
        nejdelsi = len(nazev) * 0.5
        for z in zaznamy:
            nejdelsi = max(nejdelsi, len(str(dict(z["pary"]).get(klic, ""))))
        ws.column_dimensions[pismeno].width = min(max(nejdelsi + 4, 12), 30)
    ws.freeze_panes = f"A{pr}"
    return varovani


def sestav(kolo: dict, slozka: Path) -> Path:
    wb = Workbook()
    _list_prehled(wb, kolo)
    varovani: list[str] = []
    for typ in kolo["formulare"]:
        if typ in ROZVRZENI:
            varovani += _list_bloku(wb, kolo, slozka, typ)
    for radek in varovani:
        s.vypis("  ! " + radek)
    wb.calculation.fullCalcOnLoad = True
    cil = slozka / "vysledky.xlsx"
    wb.save(cil)
    return cil


def main() -> None:
    kolo = s.nacti_kolo()
    if not kolo:
        raise SystemExit("Žádné kolo neexistuje.")
    s.vypis(f"Sešit uložen: {sestav(kolo, s.slozka_kola(kolo))}")


if __name__ == "__main__":
    main()
